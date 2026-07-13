import io
import json
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import threading
from app.database import get_db, SessionLocal
from app.tasks import create_task, run_task
from app.cache import cache
from app.models import Hospital, QualityScore, ValidationResult, AnomalyResult, IndicatorValue, ConfidenceScore
from app.schemas import ReportOut, ReportSummaryOut
from app.engine.pipeline import run_full_analysis

router = APIRouter(prefix="/reports", tags=["reports"])


@router.get("/", response_model=List[ReportOut])
def list_reports(
    month: Optional[str] = Query(None, description="Filter by month YYYY-MM"),
    source_file: Optional[str] = Query(None, description="Filter by source file name"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db),
):
    query = db.query(QualityScore)
    if month:
        query = query.filter(QualityScore.month == month)
    if source_file:
        file_subs = (
            db.query(IndicatorValue.hospital_id, IndicatorValue.month)
            .filter(IndicatorValue.source_file == source_file)
            .distinct()
            .subquery()
        )
        query = query.join(
            file_subs,
            (QualityScore.hospital_id == file_subs.c.hospital_id) & (QualityScore.month == file_subs.c.month),
        )
    cache_key = cache.make_key("reports:list", month=month, source_file=source_file, skip=skip, limit=limit)
    cached = cache.get(cache_key)
    if cached:
        return cached

    scores = query.order_by(QualityScore.month).offset(skip).limit(limit).all()
    if not scores:
        return []

    hospital_ids = list(set(s.hospital_id for s in scores))
    hospitals = {h.id: h for h in db.query(Hospital).filter(Hospital.id.in_(hospital_ids)).all()}

    # Filter out inactive hospitals
    scores = [s for s in scores if hospitals.get(s.hospital_id) and hospitals[s.hospital_id].is_active]
    if not scores:
        return []
    hospital_ids = list(set(s.hospital_id for s in scores))

    months = list(set(s.month for s in scores))
    anomaly_rows = (
        db.query(AnomalyResult)
        .filter(
            AnomalyResult.hospital_id.in_(hospital_ids),
            AnomalyResult.month.in_(months),
            AnomalyResult.is_outlier,
        )
        .all()
    )
    anomaly_by_key: dict = {}
    for a in anomaly_rows:
        key = (a.hospital_id, a.month)
        anomaly_by_key.setdefault(key, []).append(a)

    results = []
    for s in scores:
        hosp = hospitals.get(s.hospital_id)
        issues = json.loads(s.issues) if s.issues else []
        outliers = [
            {"indicator": a.rate_name, "value": a.value, "benchmark": a.benchmark}
            for a in anomaly_by_key.get((s.hospital_id, s.month), [])
        ]
        results.append(ReportOut(
            hospital=hosp.name if hosp else "Unknown",
            month=s.month,
            data_quality_score=s.score,
            rule_compliance=s.rule_compliance,
            completeness=s.completeness,
            consistency=s.consistency,
            outlier_penalty=s.outlier_penalty,
            issues=issues,
            outliers=outliers,
        ))
    cache.set(cache_key, results)
    return results


@router.get("/generate/{hospital_id}", response_model=ReportOut)
def generate_report(hospital_id: int, month: str = Query(..., description="Month YYYY-MM"), db: Session = Depends(get_db)):
    try:
        report = run_full_analysis(db, hospital_id, month)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return ReportOut(**report)


@router.post("/generate/{hospital_id}")
def generate_report_background(
    hospital_id: int,
    month: str = Query(..., description="Month YYYY-MM"),
    background_tasks: BackgroundTasks = None,
):
    task_id = create_task("Generate Report", lambda: None)

    def _run(tid, hid, m):
        bg_db = SessionLocal()
        try:
            from app.engine.pipeline import run_full_analysis
            from app.tasks import set_progress, set_status
            set_progress(tid, 50)
            run_full_analysis(bg_db, hid, m)
            set_progress(tid, 100)
            set_status(tid, "done")
        except Exception as e:
            from app.tasks import set_status
            set_status(tid, "error")
            raise e
        finally:
            bg_db.close()

    if background_tasks is not None:
        background_tasks.add_task(run_task, task_id, _run, task_id, hospital_id, month)
    else:
        threading.Thread(target=run_task, args=(task_id, _run, task_id, hospital_id, month), daemon=True).start()

    return {"task_id": task_id, "message": f"Report generation started. Use /tasks/{task_id} to check status."}


@router.get("/detail/{hospital_id}", response_model=ReportSummaryOut)
def detailed_report(hospital_id: int, month: str = Query(..., description="Month YYYY-MM"), db: Session = Depends(get_db)):
    hosp = db.query(Hospital).filter(Hospital.id == hospital_id).first()
    if not hosp:
        raise HTTPException(status_code=404, detail="Hospital not found")

    # Read stored score from database (consistent with list endpoint)
    qs = db.query(QualityScore).filter(
        QualityScore.hospital_id == hospital_id,
        QualityScore.month == month,
    ).first()
    if not qs:
        # Run analysis if no stored score exists
        try:
            report = run_full_analysis(db, hospital_id, month, force=True)
        except ValueError as e:
            raise HTTPException(status_code=404, detail=str(e))
        data_quality_score = report["data_quality_score"]
        rule_compliance = report.get("rule_compliance")
        completeness = report.get("completeness")
        consistency = report.get("consistency")
        outlier_penalty = report.get("outlier_penalty")
        issues = report["issues"]
        confidence = report.get("confidence")
    else:
        data_quality_score = qs.score
        rule_compliance = qs.rule_compliance
        completeness = qs.completeness
        consistency = qs.consistency
        outlier_penalty = qs.outlier_penalty
        issues = json.loads(qs.issues) if qs.issues else []
        # Try to read confidence from database
        cs = db.query(ConfidenceScore).filter(
            ConfidenceScore.hospital_id == hospital_id,
            ConfidenceScore.month == month,
        ).first()
        confidence = json.loads(cs.confidence_data) if cs and cs.confidence_data else None

    validation_rows = db.query(ValidationResult).filter(
        ValidationResult.hospital_id == hospital_id,
        ValidationResult.month == month,
    ).all()
    anomaly_rows = db.query(AnomalyResult).filter(
        AnomalyResult.hospital_id == hospital_id,
        AnomalyResult.month == month,
    ).all()
    return ReportSummaryOut(
        hospital=hosp.name,
        month=month,
        data_quality_score=data_quality_score,
        rule_compliance=rule_compliance,
        completeness=completeness,
        consistency=consistency,
        outlier_penalty=outlier_penalty,
        issues=issues,
        validation_results=validation_rows,
        anomaly_results=anomaly_rows,
        confidence=confidence,
    )


@router.get("/from-file/{source_file}", response_model=List[ReportOut])
def reports_from_file(source_file: str, db: Session = Depends(get_db)):
    values = (
        db.query(IndicatorValue)
        .filter(IndicatorValue.source_file == source_file)
        .all()
    )
    if not values:
        raise HTTPException(status_code=404, detail=f"No data found for file: {source_file}")

    hospital_months = set()
    for v in values:
        hospital_months.add((v.hospital_id, v.month))

    reports = []
    for hospital_id, month in sorted(hospital_months):
        try:
            report = run_full_analysis(db, hospital_id, month)
            reports.append(report)
        except Exception:
            pass
    return reports


def _get_export_data(db: Session, hospital_id: int, month: str) -> dict:
    hosp = db.query(Hospital).filter(Hospital.id == hospital_id).first()
    if not hosp:
        raise HTTPException(status_code=404, detail="Hospital not found")
    try:
        report = run_full_analysis(db, hospital_id, month)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    vrows = db.query(ValidationResult).filter(
        ValidationResult.hospital_id == hospital_id,
        ValidationResult.month == month,
    ).all()
    arows = db.query(AnomalyResult).filter(
        AnomalyResult.hospital_id == hospital_id,
        AnomalyResult.month == month,
    ).all()
    return {"hospital": hosp.name, "month": month, "report": report, "validations": vrows, "anomalies": arows}


def _esc(text: str) -> str:
    s = str(text)
    s = s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
    return "".join(ch if ord(ch) < 128 else f"&#{ord(ch)};" for ch in s)


def _register_unicode_font():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os as _os
    candidates = [
        r"C:\Windows\Fonts\segoeui.ttf",
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
    ]
    for p in candidates:
        if _os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont("UniFont", p))
                return "UniFont"
            except Exception:
                continue
    return None


@router.get("/export/{hospital_id}/pdf")
def export_report_pdf(hospital_id: int, month: str = Query(...), db: Session = Depends(get_db)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    try:
        d = _get_export_data(db, hospital_id, month)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load report data: {e}")
    try:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        story = []

        uni_name = _register_unicode_font()
        if uni_name:
            for s in styles.byName.values():
                s.fontName = uni_name

        story.append(Paragraph(f"<b>{_esc(d['hospital'])}</b>", styles["Title"]))
        story.append(Paragraph(f"Monthly Quality Report &mdash; {_esc(d['month'])}", styles["Normal"]))
        story.append(Spacer(1, 6*mm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1a237e")))
        story.append(Spacer(1, 4*mm))

        r = d["report"]
        score = r.get("data_quality_score", 0)
        score_color = colors.HexColor("#2e7d32") if score >= 80 else colors.HexColor("#e65100") if score >= 50 else colors.HexColor("#c62828")

        sc_style = ParagraphStyle("Score", parent=styles["Normal"], fontSize=28, textColor=score_color, alignment=1)
        if uni_name:
            sc_style.fontName = uni_name
        story.append(Paragraph(f"{score:.0f}/100", sc_style))
        story.append(Spacer(1, 3*mm))

        sub_data = [
            ["Metric", "Value"],
            ["Rule Compliance", f"{r.get('rule_compliance', 0):.1f}%" if r.get('rule_compliance') is not None else "--"],
            ["Completeness", f"{r.get('completeness', 0):.1f}%" if r.get('completeness') is not None else "--"],
            ["Consistency", f"{r.get('consistency', 0):.1f}%" if r.get('consistency') is not None else "--"],
            ["Outlier Penalty", f"{r.get('outlier_penalty', 0):.2f}" if r.get('outlier_penalty') is not None else "--"],
        ]
        t = Table(sub_data, colWidths=[120, 80])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (1, 1), (1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ]))
        story.append(t)
        story.append(Spacer(1, 4*mm))

        issues = r.get("issues", [])
        if issues:
            story.append(Paragraph("<b>Issues Found</b>", styles["Heading2"]))
            for iss in issues:
                story.append(Paragraph(f"&bull; {_esc(iss)}", styles["Normal"]))
            story.append(Spacer(1, 3*mm))

        vrows = d["validations"]
        story.append(Paragraph(f"<b>Validation Results ({len(vrows)})</b>", styles["Heading2"]))
        vdata = [["Code", "Description", "Status", "Severity", "Details"]]
        fails = 0
        for v in vrows:
            vdata.append([
                _esc(v.rule_code),
                _esc(v.rule_description)[:40],
                v.status,
                v.severity,
                _esc(v.details or "")[:50],
            ])
            if v.status == "FAIL":
                fails += 1
        vt = Table(vdata, colWidths=[40, 100, 40, 40, 100], repeatRows=1)
        vt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (2, 1), (3, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(vt)
        story.append(Spacer(1, 2*mm))

        arows = d["anomalies"]
        story.append(Paragraph(f"<b>Anomaly Detection ({len(arows)})</b>", styles["Heading2"]))
        adata = [["Rate", "Value", "Benchmark", "Z-Score", "Outlier"]]
        outlier_count = 0
        for a in arows:
            adata.append([
                _esc(a.rate_name),
                f"{a.value:.2f}" if a.value is not None else "--",
                f"{a.benchmark:.2f}" if a.benchmark is not None else "--",
                f"{a.z_score:.2f}" if a.z_score is not None else "--",
                "Yes" if a.is_outlier else "No",
            ])
            if a.is_outlier:
                outlier_count += 1
        at = Table(adata, colWidths=[100, 60, 60, 50, 40], repeatRows=1)
        at.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(at)
        story.append(Spacer(1, 4*mm))

        style_footer = ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=colors.grey)
        if uni_name:
            style_footer.fontName = uni_name
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
        story.append(Paragraph(f"Generated by HEALTH-ai | {datetime.now().strftime('%Y-%m-%d %H:%M')} | {fails} failures, {outlier_count} outliers", style_footer))

        doc.build(story)
        buf.seek(0)
        filename = f"report_{d['hospital']}_{d['month']}.pdf".replace(" ", "_")
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {e}")


@router.get("/export/{hospital_id}/excel")
def export_report_excel(hospital_id: int, month: str = Query(...), db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    d = _get_export_data(db, hospital_id, month)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quality Report"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = d["report"]
    hospital_name = d["hospital"]
    report_month = d["month"]
    issues = r.get("issues", [])
    vrows = d["validations"]
    arows = d["anomalies"]

    row = 1
    ws.cell(row, 1, f"Quality Report - {hospital_name}").font = Font(bold=True, size=14, color="1A237E")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    ws.cell(row, 1, report_month).font = Font(size=11, color="888888")
    row += 2

    ws.cell(row, 1, "Metric").font = hdr_font
    ws.cell(row, 1).fill = hdr_fill
    ws.cell(row, 2, "Value").font = hdr_font
    ws.cell(row, 2).fill = hdr_fill
    for c in [ws.cell(row, 1), ws.cell(row, 2)]:
        c.border = border
        c.alignment = Alignment(horizontal="center")
    row += 1
    for label, key in [("Data Quality Score", "data_quality_score"), ("Rule Compliance", "rule_compliance"),
                        ("Completeness", "completeness"), ("Consistency", "consistency"),
                        ("Outlier Penalty", "outlier_penalty")]:
        val = r.get(key)
        if val is not None:
            disp = f"{val:.1f}%" if key != "data_quality_score" else f"{val:.2f}"
            if key == "data_quality_score":
                disp = f"{val:.0f}/100"
        else:
            disp = "--"
        ws.cell(row, 1, label).border = border
        ws.cell(row, 2, disp).border = border
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    # Issues
    ws.cell(row, 1, "Issues Found").font = hdr_font
    ws.cell(row, 1).fill = hdr_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1).border = border
    row += 1
    if issues:
        for iss in issues:
            ws.cell(row, 1, iss).border = border
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1
    else:
        ws.cell(row, 1, "No issues found").border = border
        row += 1

    row += 1
    # Validation Results
    ws.cell(row, 1, "Validation Results").font = hdr_font
    ws.cell(row, 1).fill = hdr_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1).border = border
    row += 1
    for col, h in enumerate(["Code", "Description", "Status", "Severity", "Details"], 1):
        c = ws.cell(row, col, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    row += 1
    for v in vrows:
        ws.cell(row, 1, v.rule_code).border = border
        ws.cell(row, 2, v.rule_description).border = border
        sc = ws.cell(row, 3, v.status)
        sc.border = border
        sc.alignment = Alignment(horizontal="center")
        if v.status == "FAIL":
            sc.font = Font(color="C62828", bold=True)
        else:
            sc.font = Font(color="2E7D32")
        ws.cell(row, 4, v.severity).border = border
        ws.cell(row, 4).alignment = Alignment(horizontal="center")
        ws.cell(row, 5, v.details or "").border = border
        row += 1

    row += 1
    # Anomalies
    ws.cell(row, 1, "Anomaly Detection").font = hdr_font
    ws.cell(row, 1).fill = hdr_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1).border = border
    row += 1
    for col, h in enumerate(["Rate Name", "Value", "Benchmark", "Z-Score", "Outlier"], 1):
        c = ws.cell(row, col, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    row += 1
    for a in arows:
        ws.cell(row, 1, a.rate_name).border = border
        ws.cell(row, 2, f"{a.value:.2f}" if a.value is not None else "--").border = border
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
        ws.cell(row, 3, f"{a.benchmark:.2f}" if a.benchmark is not None else "--").border = border
        ws.cell(row, 3).alignment = Alignment(horizontal="center")
        ws.cell(row, 4, f"{a.z_score:.2f}" if a.z_score is not None else "--").border = border
        ws.cell(row, 4).alignment = Alignment(horizontal="center")
        oc = ws.cell(row, 5, "Yes" if a.is_outlier else "No")
        oc.border = border
        oc.alignment = Alignment(horizontal="center")
        if a.is_outlier:
            oc.font = Font(color="C62828", bold=True)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"report_{hospital_name}_{report_month}.xlsx".replace(" ", "_")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})